from datetime import datetime, timedelta

import openpyxl
from django import forms
from django.contrib import admin
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from agents.constant import COMPLETED_PERCENTAGE, MAX_MINUTES_IN_STORE, MAX_VISIBLE_AGENTS

from .models import DailyPlan, PhotoPost


class DateRangeForm(forms.Form):
    start_date = forms.DateField(
        widget=forms.DateInput(attrs={"type": "date", "name": "start_date"}),
        required=False,
        label="С даты (День/Месяц/Год)",
    )
    end_date = forms.DateField(
        widget=forms.DateInput(attrs={"type": "date", "name": "end_date"}),
        required=False,
        label="По дату (День/Месяц/Год)",
    )


def export_to_excel(modeladmin, request, queryset):
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")

    if "apply" not in request.POST:
        form = DateRangeForm()
        context = {
            "queryset": queryset,
            "form": form,
            "action_checkbox_name": admin.helpers.ACTION_CHECKBOX_NAME,
        }
        return render(request, "export_photo_report.html", context)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет фотопостов"

    if start_date and end_date:
        start_date_parsed = datetime.strptime(start_date, "%Y-%m-%d")
        end_date_parsed = datetime.strptime(end_date, "%Y-%m-%d")
    elif start_date:
        start_date_parsed = datetime.strptime(start_date, "%Y-%m-%d")
        end_date_parsed = start_date_parsed + timedelta(days=6)
    elif end_date:
        end_date_parsed = datetime.strptime(end_date, "%Y-%m-%d")
        start_date_parsed = end_date_parsed - timedelta(days=6)
    else:
        end_date_parsed = datetime.now()
        start_date_parsed = end_date_parsed - timedelta(days=6)

    selected_agents = queryset

    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    link_font = Font(color="0000FF", underline="single")

    dmp_orimi_brands = ["Гринф", "Tess", "ЖН", "Шах", "Jardin", "Жокей"]
    dmp_competitor_brands = ["Бета", "Пиала", "Ахмад", "Jacobs", "Nestlе"]
    all_dmp_brands = dmp_orimi_brands + dmp_competitor_brands

    brand_mapping = {
        "гринф": ["гринф", "greenf", "grinf"],
        "tess": ["tess", "тесс"],
        "жн": ["жн", "jn"],
        "шах": ["шах", "shah"],
        "jardin": ["jardin", "жардин"],
        "жокей": ["жокей", "jockey"],
        "бета": ["бета", "beta"],
        "пиала": ["пиала", "piala"],
        "ахмад": ["ахмад", "ahmad"],
        "jacobs": ["jacobs", "якобс"],
        "nestlе": ["nestle", "нестле", "nestlе"],
    }

    def check_brand_match(dmp_type, brand, brand_variants):
        if not dmp_type:
            return False
        dmp_type_lower = dmp_type.strip().lower()

        if dmp_type_lower == brand.lower():
            return True

        return dmp_type_lower in brand_variants

    row = 1

    main_headers = [
        "Дата",
        "SNO",
        "Наименование TT",
        "Начала работы в TT",
        "Завершение работы в TT",
        "Time in outlet",
        "РМП_чай_ДО",
        "РМП_чай_ПОСЛЕ",
        "РМП_кофе_ДО",
        "РМП_кофе_ПОСЛЕ",
        "ДМП_ОРИМИ КР",
        "ДМП_конкурент",
    ]

    col = 1
    for i, header in enumerate(main_headers[:10]):
        ws.cell(row=row, column=col + i, value=header).font = header_font
        ws.merge_cells(start_row=row, start_column=col + i, end_row=row + 1, end_column=col + i)
        ws.cell(row=row, column=col + i).alignment = center_align

    col_offset = 11
    ws.cell(row=row, column=col_offset, value="ДМП_ОРИМИ КР").font = header_font
    ws.merge_cells(
        start_row=row,
        start_column=col_offset,
        end_row=row,
        end_column=col_offset + len(dmp_orimi_brands) - 1,
    )
    ws.cell(row=row, column=col_offset).alignment = center_align

    col_offset += len(dmp_orimi_brands)
    ws.cell(row=row, column=col_offset, value="ДМП_конкурент").font = header_font
    ws.merge_cells(
        start_row=row,
        start_column=col_offset,
        end_row=row,
        end_column=col_offset + len(dmp_competitor_brands) - 1,
    )
    ws.cell(row=row, column=col_offset).alignment = center_align

    col = 11
    for brand in all_dmp_brands:
        ws.cell(row=row + 1, column=col, value=brand).font = header_font
        ws.cell(row=row + 1, column=col).alignment = center_align
        col += 1

    data_row = 3

    total_days = (end_date_parsed - start_date_parsed).days + 1

    for agent in selected_agents:
        agent_has_data = False

        for day_offset in range(total_days):
            current_date = start_date_parsed + timedelta(days=day_offset)
            date_str = current_date.strftime("%d.%m.%Y")

            day_posts = (
                PhotoPost.objects.filter(agent=agent, created__date=current_date.date())
                .select_related("store")
                .order_by("created")
            )

            if not day_posts.exists():
                continue

            agent_has_data = True

            date_start_row = data_row

            store_groups = []
            current_store = None
            current_group = []

            for post in day_posts:
                if post.store != current_store:
                    if current_group:
                        store_groups.append((current_store, current_group))
                    current_store = post.store
                    current_group = [post]
                else:
                    current_group.append(post)

            if current_group:
                store_groups.append((current_store, current_group))

            first_store_in_day = True

            for store, posts_list in store_groups:
                rmp_photos = {}
                for post_type in [
                    "РМП_чай_ДО",
                    "РМП_чай_ПОСЛЕ",
                    "РМП_кофе_ДО",
                    "РМП_кофе_ПОСЛЕ",
                ]:
                    type_posts = [p for p in posts_list if p.post_type == post_type and p.image]
                    if type_posts:
                        rmp_photos[post_type] = type_posts

                dmp_orimi_photos = {}
                dmp_orimi_posts = [p for p in posts_list if p.post_type == "ДМП_ОРИМИ КР" and p.image]
                for brand in dmp_orimi_brands:
                    brand_variants = brand_mapping.get(brand.lower(), [brand.lower()])
                    brand_posts = []
                    for post in dmp_orimi_posts:
                        if check_brand_match(post.dmp_type, brand, brand_variants):
                            brand_posts.append(post)
                    if brand_posts:
                        dmp_orimi_photos[brand] = brand_posts

                max_rmp_photos = max(len(photos) for photos in rmp_photos.values()) if rmp_photos else 0
                max_dmp_photos = max(len(photos) for photos in dmp_orimi_photos.values()) if dmp_orimi_photos else 0
                max_photos = max(max_rmp_photos, max_dmp_photos, 1)

                for photo_index in range(max_photos):
                    if photo_index == 0:
                        if first_store_in_day:
                            ws.cell(row=data_row, column=1, value=date_str)
                            ws.cell(row=data_row, column=2, value=agent.agent_name)
                            first_store_in_day = False

                        ws.cell(row=data_row, column=3, value=store.name)

                        first_post = min(posts_list, key=lambda x: x.created)
                        last_post = max(posts_list, key=lambda x: x.created)
                        first_created = first_post.created + timedelta(hours=6)
                        last_created = last_post.created + timedelta(hours=6)
                        ws.cell(
                            row=data_row,
                            column=4,
                            value=first_created.strftime("%H:%M"),
                        )
                        ws.cell(
                            row=data_row,
                            column=5,
                            value=last_created.strftime("%H:%M"),
                        )

                        time_diff = last_post.created - first_post.created
                        hours = int(time_diff.total_seconds() // 3600)
                        minutes = int((time_diff.total_seconds() % 3600) // 60)
                        ws.cell(row=data_row, column=6, value=f"{hours}:{minutes:02d}")

                    col = 7
                    for post_type in [
                        "РМП_чай_ДО",
                        "РМП_чай_ПОСЛЕ",
                        "РМП_кофе_ДО",
                        "РМП_кофе_ПОСЛЕ",
                    ]:
                        if post_type in rmp_photos and photo_index < len(rmp_photos[post_type]):
                            post = rmp_photos[post_type][photo_index]
                            image_url = request.build_absolute_uri(post.image.url)
                            ws.cell(
                                row=data_row,
                                column=col,
                                value=f'=HYPERLINK("{image_url}","фото")',
                            )
                            ws.cell(row=data_row, column=col).font = link_font
                        else:
                            ws.cell(row=data_row, column=col, value="")
                        col += 1

                    col = 11
                    for brand in dmp_orimi_brands:
                        if brand in dmp_orimi_photos and photo_index < len(dmp_orimi_photos[brand]):
                            post = dmp_orimi_photos[brand][photo_index]
                            image_url = request.build_absolute_uri(post.image.url)
                            ws.cell(
                                row=data_row,
                                column=col,
                                value=f'=HYPERLINK("{image_url}","фото")',
                            )
                            ws.cell(row=data_row, column=col).font = link_font
                        else:
                            ws.cell(row=data_row, column=col, value="")
                        col += 1

                    if photo_index == 0:
                        dmp_competitor_posts = [p for p in posts_list if p.post_type == "ДМП_конкурент"]
                        for brand in dmp_competitor_brands:
                            brand_variants = brand_mapping.get(brand.lower(), [brand.lower()])

                            brand_posts = []
                            for post in dmp_competitor_posts:
                                if check_brand_match(post.dmp_type, brand, brand_variants):
                                    brand_posts.append(post)

                            if brand_posts:
                                total_count = sum(p.dmp_count or 1 for p in brand_posts)
                                ws.cell(row=data_row, column=col, value=total_count)
                            else:
                                ws.cell(row=data_row, column=col, value="")
                            col += 1

                    data_row += 1

            if date_start_row < data_row:
                ws.merge_cells(
                    start_row=date_start_row,
                    start_column=1,
                    end_row=data_row - 1,
                    end_column=1,
                )
                ws.merge_cells(
                    start_row=date_start_row,
                    start_column=2,
                    end_row=data_row - 1,
                    end_column=2,
                )

        if agent_has_data:
            data_row += 1

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    column_widths = {
        1: 12,
        2: 15,
        3: 25,
        4: 15,
        5: 15,
        6: 12,
        7: 15,
        8: 15,
        9: 15,
        10: 15,
    }

    for col_num, width in column_widths.items():
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = width

    for col_num in range(11, 11 + len(all_dmp_brands)):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = 10

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    agent_names = "_".join([agent.agent_name[:10].replace(" ", "_") for agent in selected_agents[:3]])
    if len(selected_agents) > MAX_VISIBLE_AGENTS:
        agent_names += f"_and_{len(selected_agents) - 3}_more"

    date_suffix = ""
    if start_date and end_date:
        date_suffix = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_suffix = f"_from_{start_date}"
    elif end_date:
        date_suffix = f"_to_{end_date}"

    filename = f"photo_report_{agent_names}{date_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"

    wb.save(response)

    filter_info = ""
    if start_date or end_date:
        filter_info = f" (период: {start_date or 'начало'} - {end_date or 'конец'})"

    modeladmin.message_user(request, f"Отчет успешно создан для {len(selected_agents)} агент(ов){filter_info}")

    return response


export_to_excel.short_description = "Выгрузить отчет фотопостов в Excel"


def _get_excel_styles():
    header_font = Font(bold=True, size=10)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    success_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    fail_fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    return header_font, center_align, thin_border, success_fill, fail_fill, header_fill


def _setup_excel_headers(ws, header_font, header_fill, all_dmp_brands):
    headers_row1 = [
        ("Дата", 1, 1),
        ("Мерчендайзер", 2, 2),
        ("План визита", 3, 3),
        ("Факт визита", 4, 4),
        ("Call completion %", 5, 5),
        ("Кол-во Визитов < 30 мин", 6, 6),
        ("Кол-во Визитов > 30 мин", 7, 7),
        ("Общая Время в ТТ", 8, 8),
        ("РМП_чай", 9, 10),
        ("РМП_кофе", 11, 12),
        ("ДМП ОРИМИ", 13, 14),
    ]

    current_col = 15
    for brand in all_dmp_brands:
        headers_row1.append((brand, current_col, current_col))
        current_col += 1

    headers_row2 = [
        (9, "Кол-во ТТ"),
        (10, "кол-во фото"),
        (11, "Кол-во ТТ"),
        (12, "кол-во фото"),
        (13, "Кол-во ТТ"),
        (14, "сумма"),
    ]

    for header, start_col, end_col in headers_row1:
        ws.cell(row=1, column=start_col, value=header).font = header_font
        ws.cell(row=1, column=start_col).fill = header_fill
        if start_col != end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            ws.cell(row=1, column=end_col).fill = header_fill

    for col, header in headers_row2:
        ws.cell(row=2, column=col, value=header).font = header_font
        ws.cell(row=2, column=col).fill = header_fill

    single_level_headers = list(range(1, 9)) + list(range(15, 15 + len(all_dmp_brands)))
    for i in single_level_headers:
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)


def _calculate_visit_metrics(agent, date):
    visits_under_30 = 0
    visits_over_30 = 0
    total_time_minutes = 0

    visited_stores = (
        PhotoPost.objects.filter(agent=agent, created__date=date).values_list("store", flat=True).distinct()
    )

    for store_id in visited_stores:
        store_posts = PhotoPost.objects.filter(agent=agent, store_id=store_id, created__date=date).order_by(
            "created"
        )

        if store_posts.exists():
            first_post = store_posts.first()
            last_post = store_posts.last()
            time_diff = last_post.created - first_post.created
            minutes_in_store = int(time_diff.total_seconds() // 60)

            if minutes_in_store < MAX_MINUTES_IN_STORE:
                visits_under_30 += 1
            else:
                visits_over_30 += 1

            total_time_minutes += minutes_in_store

    hours = total_time_minutes // 60
    minutes = total_time_minutes % 60
    total_time_str = f"{hours}:{minutes:02d}"

    return visits_under_30, visits_over_30, total_time_str


def _process_posts_data(agent, date, all_dmp_brands, dmp_orimi_brands, dmp_competitor_brands):
    posts_data = PhotoPost.objects.filter(agent=agent, created__date=date).values(
        "post_type", "store", "dmp_type", "dmp_count"
    )

    rmp_tea_stores = set()
    rmp_tea_photos = 0
    rmp_coffee_stores = set()
    rmp_coffee_photos = 0
    dmp_orimi_stores = set()
    dmp_orimi_sum = 0

    brand_sums = dict.fromkeys(all_dmp_brands, 0)
    brand_photo_counts = dict.fromkeys(all_dmp_brands, 0)

    for post in posts_data:
        post_type = post["post_type"]
        store_id = post["store"]
        dmp_type = post["dmp_type"]
        dmp_count = post["dmp_count"] or 0

        if "РМП_чай" in post_type:
            rmp_tea_stores.add(store_id)
            rmp_tea_photos += 1
        elif "РМП_кофе" in post_type:
            rmp_coffee_stores.add(store_id)
            rmp_coffee_photos += 1
        elif "ДМП_ОРИМИ" in post_type:
            dmp_orimi_stores.add(store_id)
            dmp_orimi_sum += 1

            brand = get_brand_from_dmp_type(dmp_type)
            if brand and brand in dmp_orimi_brands:
                brand_photo_counts[brand] += 1

        elif "ДМП_конкурент" in post_type:
            if dmp_count > 0:
                brand = get_brand_from_dmp_type(dmp_type)
                if brand and brand in dmp_competitor_brands:
                    brand_sums[brand] += dmp_count

    return (rmp_tea_stores, rmp_tea_photos, rmp_coffee_stores, rmp_coffee_photos,
            dmp_orimi_stores, dmp_orimi_sum, brand_sums, brand_photo_counts)


def _write_row_data(ws, row, date, agent, plan_count, fact_count, visits_under_30, visits_over_30,
                   total_time_str, rmp_tea_stores, rmp_tea_photos, rmp_coffee_stores, rmp_coffee_photos,
                   dmp_orimi_stores, dmp_orimi_sum, brand_sums, brand_photo_counts, all_dmp_brands,
                   dmp_orimi_brands, success_fill, fail_fill):
    ws.cell(row=row, column=1, value=date.strftime("%d.%m"))
    ws.cell(row=row, column=2, value=agent.agent_name)
    ws.cell(row=row, column=3, value=plan_count)
    ws.cell(row=row, column=4, value=fact_count)

    if plan_count > 0:
        completion_percent = round((fact_count / plan_count) * 100)
        ws.cell(row=row, column=5, value=f"{completion_percent}%")
        if completion_percent >= COMPLETED_PERCENTAGE:
            ws.cell(row=row, column=5).fill = success_fill
        else:
            ws.cell(row=row, column=5).fill = fail_fill
    else:
        ws.cell(row=row, column=5, value="0%")
        ws.cell(row=row, column=5).fill = PatternFill(
            start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
        )

    ws.cell(row=row, column=6, value=visits_under_30)
    ws.cell(row=row, column=7, value=visits_over_30)
    ws.cell(row=row, column=8, value=total_time_str)
    ws.cell(row=row, column=9, value=len(rmp_tea_stores))
    ws.cell(row=row, column=10, value=rmp_tea_photos)
    ws.cell(row=row, column=11, value=len(rmp_coffee_stores))
    ws.cell(row=row, column=12, value=rmp_coffee_photos)
    ws.cell(row=row, column=13, value=len(dmp_orimi_stores))
    ws.cell(row=row, column=14, value=dmp_orimi_sum)

    col = 15
    for brand in all_dmp_brands:
        if brand in dmp_orimi_brands:
            ws.cell(row=row, column=col, value=brand_photo_counts[brand])
        else:
            ws.cell(row=row, column=col, value=brand_sums[brand])
        col += 1


def _apply_excel_formatting(ws, center_align, thin_border, all_dmp_brands):
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    column_widths = {
        "A": 12, "B": 20, "C": 12, "D": 12, "E": 15, "F": 18, "G": 18, "H": 15,
        "I": 10, "J": 12, "K": 10, "L": 12, "M": 12, "N": 12,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    for i in range(15, 15 + len(all_dmp_brands)):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 10


def export_plan_visits_to_excel(modeladmin, request, queryset):
    start_date = request.POST.get("start_date")
    end_date = request.POST.get("end_date")

    if "apply" not in request.POST:
        form = DateRangeForm()
        context = {
            "queryset": queryset,
            "form": form,
            "action_checkbox_name": admin.helpers.ACTION_CHECKBOX_NAME,
        }
        return render(request, "export_plan_visits.html", context)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "План визитов"

    header_font, center_align, thin_border, success_fill, fail_fill, header_fill = _get_excel_styles()

    dmp_orimi_brands = ["Гринф", "Tess", "ЖН", "Шах", "Jardin", "Жокей"]
    dmp_competitor_brands = ["Beta", "Пиала", "Ахмад", "Jacobs", "Nestle"]
    all_dmp_brands = dmp_orimi_brands + dmp_competitor_brands

    def get_brand_from_dmp_type(dmp_type):
        if not dmp_type:
            return None

        dmp_type_lower = dmp_type.strip().lower()

        brand_mapping = {
            "tess": "Tess",
            "тесс": "Tess",
            "гринф": "Гринф",
            "greenf": "Гринф",
            "grinf": "Гринф",
            "жн": "ЖН",
            "jn": "ЖН",
            "шах": "Шах",
            "shah": "Шах",
            "jardin": "Jardin",
            "жардин": "Jardin",
            "жокей": "Жокей",
            "jockey": "Жокей",
            "beta": "Beta",
            "бета": "Beta",
            "пиала": "Пиала",
            "piala": "Пиала",
            "ахмад": "Ахмад",
            "ahmad": "Ахмад",
            "jacobs": "Jacobs",
            "якобс": "Jacobs",
            "nestle": "Nestle",
            "нестле": "Nestle",
        }

        if dmp_type_lower in brand_mapping:
            return brand_mapping[dmp_type_lower]

        for brand in all_dmp_brands:
            if brand.lower() == dmp_type_lower:
                return brand

        return None

    headers_row1 = [
        ("Дата", 1, 1),
        ("Мерчендайзер", 2, 2),
        ("План визита", 3, 3),
        ("Факт визита", 4, 4),
        ("Call completion %", 5, 5),
        ("Кол-во Визитов < 30 мин", 6, 6),
        ("Кол-во Визитов > 30 мин", 7, 7),
        ("Общая Время в ТТ", 8, 8),
        ("РМП_чай", 9, 10),
        ("РМП_кофе", 11, 12),
        ("ДМП ОРИМИ", 13, 14),
    ]

    current_col = 15
    for brand in all_dmp_brands:
        headers_row1.append((brand, current_col, current_col))
        current_col += 1

    headers_row2 = [
        (9, "Кол-во ТТ"),
        (10, "кол-во фото"),
        (11, "Кол-во ТТ"),
        (12, "кол-во фото"),
        (13, "Кол-во ТТ"),
        (14, "сумма"),
    ]

    for header, start_col, end_col in headers_row1:
        ws.cell(row=1, column=start_col, value=header).font = header_font
        ws.cell(row=1, column=start_col).fill = header_fill
        if start_col != end_col:
            ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            ws.cell(row=1, column=end_col).fill = header_fill

    for col, header in headers_row2:
        ws.cell(row=2, column=col, value=header).font = header_font
        ws.cell(row=2, column=col).fill = header_fill

    single_level_headers = list(range(1, 9)) + list(range(15, 15 + len(all_dmp_brands)))
    for i in single_level_headers:
        ws.merge_cells(start_row=1, start_column=i, end_row=2, end_column=i)

    row = 3

    for agent in selected_agents:
        daily_plans = DailyPlan.objects.filter(agent=agent)

        if start_date:
            start_date_parsed = datetime.strptime(start_date, "%Y-%m-%d").date()
            daily_plans = daily_plans.filter(date__gte=start_date_parsed)
        if end_date:
            end_date_parsed = datetime.strptime(end_date, "%Y-%m-%d").date()
            daily_plans = daily_plans.filter(date__lte=end_date_parsed)

        daily_plans = daily_plans.order_by("date")

        for daily_plan in daily_plans:
            date = daily_plan.date
            plan_count = daily_plan.planned_stores_count
            fact_count = daily_plan.visited_stores_count

            visits_under_30 = 0
            visits_over_30 = 0
            total_time_minutes = 0

            visited_stores = (
                PhotoPost.objects.filter(agent=agent, created__date=date).values_list("store", flat=True).distinct()
            )

            for store_id in visited_stores:
                store_posts = PhotoPost.objects.filter(agent=agent, store_id=store_id, created__date=date).order_by(
                    "created"
                )

                if store_posts.exists():
                    first_post = store_posts.first()
                    last_post = store_posts.last()
                    time_diff = last_post.created - first_post.created
                    minutes_in_store = int(time_diff.total_seconds() // 60)

                    if minutes_in_store < MAX_MINUTES_IN_STORE:
                        visits_under_30 += 1
                    else:
                        visits_over_30 += 1

                    total_time_minutes += minutes_in_store

            hours = total_time_minutes // 60
            minutes = total_time_minutes % 60
            total_time_str = f"{hours}:{minutes:02d}"

            posts_data = PhotoPost.objects.filter(agent=agent, created__date=date).values(
                "post_type", "store", "dmp_type", "dmp_count"
            )

            rmp_tea_stores = set()
            rmp_tea_photos = 0
            rmp_coffee_stores = set()
            rmp_coffee_photos = 0
            dmp_orimi_stores = set()
            dmp_orimi_sum = 0
            dmp_competitor_stores = set()
            dmp_competitor_sum = 0

            brand_sums = dict.fromkeys(all_dmp_brands, 0)
            brand_photo_counts = dict.fromkeys(all_dmp_brands, 0)

            for post in posts_data:
                post_type = post["post_type"]
                store_id = post["store"]
                dmp_type = post["dmp_type"]
                dmp_count = post["dmp_count"] or 0

                if "РМП_чай" in post_type:
                    rmp_tea_stores.add(store_id)
                    rmp_tea_photos += 1
                elif "РМП_кофе" in post_type:
                    rmp_coffee_stores.add(store_id)
                    rmp_coffee_photos += 1
                elif "ДМП_ОРИМИ" in post_type:
                    dmp_orimi_stores.add(store_id)
                    dmp_orimi_sum += 1

                    brand = get_brand_from_dmp_type(dmp_type)
                    if brand and brand in dmp_orimi_brands:
                        brand_photo_counts[brand] += 1

                elif "ДМП_конкурент" in post_type:
                    dmp_competitor_stores.add(store_id)
                    if dmp_count > 0:
                        dmp_competitor_sum += dmp_count

                    brand = get_brand_from_dmp_type(dmp_type)
                    if brand and brand in dmp_competitor_brands and dmp_count > 0:
                        brand_sums[brand] += dmp_count

            ws.cell(row=row, column=1, value=date.strftime("%d.%m"))
            ws.cell(row=row, column=2, value=agent.agent_name)
            ws.cell(row=row, column=3, value=plan_count)
            ws.cell(row=row, column=4, value=fact_count)

            if plan_count > 0:
                completion_percent = round((fact_count / plan_count) * 100)
                ws.cell(row=row, column=5, value=f"{completion_percent}%")
                if completion_percent >= COMPLETED_PERCENTAGE:
                    ws.cell(row=row, column=5).fill = success_fill
                else:
                    ws.cell(row=row, column=5).fill = fail_fill
            else:
                ws.cell(row=row, column=5, value="0%")
                ws.cell(row=row, column=5).fill = PatternFill(
                    start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                )

            ws.cell(row=row, column=6, value=visits_under_30)
            ws.cell(row=row, column=7, value=visits_over_30)
            ws.cell(row=row, column=8, value=total_time_str)

            ws.cell(row=row, column=9, value=len(rmp_tea_stores))
            ws.cell(row=row, column=10, value=rmp_tea_photos)
            ws.cell(row=row, column=11, value=len(rmp_coffee_stores))
            ws.cell(row=row, column=12, value=rmp_coffee_photos)
            ws.cell(row=row, column=13, value=len(dmp_orimi_stores))
            ws.cell(row=row, column=14, value=dmp_orimi_sum)

            col = 15
            for brand in all_dmp_brands:
                if brand in dmp_orimi_brands:
                    ws.cell(row=row, column=col, value=brand_photo_counts[brand])
                else:
                    ws.cell(row=row, column=col, value=brand_sums[brand])
                col += 1

            row += 1

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = center_align
            cell.border = thin_border

    column_widths = {
        "A": 12,
        "B": 20,
        "C": 12,
        "D": 12,
        "E": 15,
        "F": 18,
        "G": 18,
        "H": 15,
        "I": 10,
        "J": 12,
        "K": 10,
        "L": 12,
        "M": 12,
        "N": 12,
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    for i in range(15, 15 + len(all_dmp_brands)):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 10

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    date_suffix = ""
    if start_date and end_date:
        date_suffix = f"_{start_date}_to_{end_date}"
    elif start_date:
        date_suffix = f"_from_{start_date}"
    elif end_date:
        date_suffix = f"_to_{end_date}"

    filename = f"plan_visits{date_suffix}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response["Content-Disposition"] = f"attachment; filename={filename}"

    wb.save(response)

    filter_info = ""
    if start_date or end_date:
        filter_info = f" (фильтр: {start_date or 'начало'} - {end_date or 'конец'})"

    modeladmin.message_user(
        request,
        f"Отчет по плану визитов создан для {len(selected_agents)} агент(ов){filter_info}",
    )

    return response


export_plan_visits_to_excel.short_description = "Выгрузить план визитов в Excel"
